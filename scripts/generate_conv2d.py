import math
import openpyxl
#wb=openpyxl.load_workbook("/home/huiwu1/workspace/int8/mkl-dnn/test_file_wh.xlsx")
#f=open(r"shape_conv_mkldnn","a+")
#wb=openpyxl.load_workbook("/home/huiwu1/workspace/int8/mkl-dnn/test_file_dw.xlsx")
#f=open(r"shape_dw_mkldnn","a+")
wb=openpyxl.load_workbook("./test_file_conv2d.xlsx")
f=open(r"../tests/benchdnn/inputs/conv/shape_conv2d_mkldnn","a+")

ws=wb.active
max_row = ws.max_row
max_col = ws.max_column 
for i in range(2,max_row):
    #for j in range(max_col):
    mb = ws.cell(row=i, column=1).value
    ic = ws.cell(row=i, column=2).value
    oc = ws.cell(row=i, column=3).value
    ih = ws.cell(row=i, column=4).value
    iw = ws.cell(row=i, column=5).value
    g = ws.cell(row=i, column=6).value
    kh = ws.cell(row=i, column=7).value
    kw = ws.cell(row=i, column=8).value
    stride_h = ws.cell(row=i, column=9).value
    stride_w = ws.cell(row=i, column=10).value
    pad_h_top = ws.cell(row=i, column=11).value
    pad_w_left = ws.cell(row=i, column=12).value
    pad_h_bottom = ws.cell(row=i, column=13).value
    pad_w_right = ws.cell(row=i, column=14).value
    d_h = ws.cell(row=i, column=15).value
    d_w = ws.cell(row=i, column=16).value
 
    if not ih:
        continue

    if not pad_h_top:
        pad_h_top = 0
    if not pad_w_left:
        pad_w_left = 0
    if not pad_h_bottom:
        pad_h_bottom = 0
    if not pad_w_right:
        pad_w_right = 0
    if not d_h:
        d_h = 0
    if not d_w:
        d_w = 0

    kh_extent = d_h * (kh - 1) + 1
    kw_extent = d_w * (kw - 1) + 1
    oh = int(((ih + pad_h_top + pad_h_bottom - kh_extent) / stride_h + 1))
    ow = int(((iw + pad_w_left + pad_w_right - kw_extent) / stride_w + 1))
    #mkldnn_str = g1mb50ic3ih224iw224oc64oh112ow112kh7kw7sh2sw2ph3pw3n"resnet_50" 
    newline = "g"+ str(g) + "mb" + str(mb) + \
	"ic" + str(ic) + "ih" + str(ih) + "iw" + str(iw) + \
	"oc" + str(oc) + "oh" + str(oh) + "ow" + str(ow) + \
	"kh" + str(kh) + "kw" + str(kw) + \
	"sh" + str(stride_h) + "sw" + str(stride_w) + \
	"ph" + str(pad_h_top) + "pw" + str(pad_w_left) + \
	"dh" + str(d_h) + "dw" + str(d_w) + \
	"n"
    
    f.writelines(newline+"\n")

f.close()
