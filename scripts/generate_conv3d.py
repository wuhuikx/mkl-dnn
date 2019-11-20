import math
import openpyxl
wb=openpyxl.load_workbook("test_file_conv3d.xlsx")
f=open(r"../tests/benchdnn/inputs/conv/shape_conv3d_mkldnn","w")
#wb=openpyxl.load_workbook("/home/huiwu1/workspace/int8/mkl-dnn/test_file_dw.xlsx")
#f=open(r"shape_conv_dw","a+")

ws=wb.active
max_row = ws.max_row
max_col = ws.max_column 
for i in range(1,max_row+1):
    #for j in range(max_col):
    mb = ws.cell(row=i, column=1).value
    ic = ws.cell(row=i, column=2).value
    oc = ws.cell(row=i, column=3).value
    it = ws.cell(row=i, column=4).value
    ih = ws.cell(row=i, column=5).value
    iw = ws.cell(row=i, column=6).value

    g = ws.cell(row=i, column=7).value

    kt = ws.cell(row=i, column=8).value
    kh = ws.cell(row=i, column=9).value
    kw = ws.cell(row=i, column=10).value

    stride_t = ws.cell(row=i, column=11).value
    stride_h = ws.cell(row=i, column=12).value
    stride_w = ws.cell(row=i, column=13).value
    
    pad_prev = ws.cell(row=i, column=14).value
    pad_h_top = ws.cell(row=i, column=15).value
    pad_w_left = ws.cell(row=i, column=16).value
    pad_next = ws.cell(row=i, column=17).value
    pad_h_bottom = ws.cell(row=i, column=18).value
    pad_w_right = ws.cell(row=i, column=19).value

    d_t = ws.cell(row=i, column=20).value
    d_h = ws.cell(row=i, column=21).value
    d_w = ws.cell(row=i, column=22).value
 
    if not ih:
        continue

    if not pad_prev:
        pad_prev = 0
    if not pad_h_top:
        pad_h_top = 0
    if not pad_w_left:
        pad_w_left = 0
    if not pad_next:
        pad_next = 0
    if not pad_h_bottom:
        pad_h_bottom = 0
    if not pad_w_right:
        pad_w_right = 0
    if not d_t:
        d_t = 1
    if not d_h:
        d_h = 1
    if not d_w:
        d_w = 1
    kt_extent = d_t * (kt - 1) + 1 
    kh_extent = d_h * (kh - 1) + 1 
    kw_extent = d_w * (kw - 1) + 1 
    ot = int(((it + pad_prev + pad_next - kt_extent) / stride_t + 1))
    oh = int(((ih + pad_h_top + pad_h_bottom - kh_extent) / stride_h + 1))
    ow = int(((iw + pad_w_left + pad_w_right - kw_extent) / stride_w + 1))
    #mb1_g1ic16oc16_kd1kh1kw1_id2ih1iw1_od1oh1ow1_sd2sh1sw1_pd0ph0pw0_dd0dh0dw0n"conv_3d_1x1_strided_no-padding_1_pd-back"
    newline = "mb" + str(mb) + \
        "_g"+ str(g) + "ic" + str(ic) + "oc" + str(oc) + \
        "_kd" + str(kt) +  "kh" + str(kh) + "kw" + str(kw) + \
        "_id" + str(it) + "ih" + str(ih) + "iw" + str(iw) + \
        "_od" + str(ot) + "oh" + str(oh) + "ow" + str(ow) + \
        "_sd" + str(stride_t) + "sh" + str(stride_h) + "sw" + str(stride_w) + \
        "_pd" + str(pad_prev) + "ph" + str(pad_h_top) + "pw" + str(pad_w_left) + \
        "_dd" + str((d_t-1)) + "dh" + str((d_h-1)) + "dw" + str((d_w-1)) + \
        "n"

    f.writelines(newline+"\n")

f.close()
